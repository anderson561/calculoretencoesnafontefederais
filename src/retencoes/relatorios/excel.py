"""Exportação dos relatórios 1, 2 e 3 para um arquivo Excel (.xlsx)."""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models import Cabecalho, NotaCalculada
from .agregacao import agregar_analitico, agregar_sintetico, agregar_trimestral
from .formatacao import NAO_INFORMADO, linhas_cabecalho, mascarar_documento

_FMT_MOEDA = 'R$ #,##0.00'
_AZUL = "1F4E78"
_CINZA = "D9D9D9"
_CINZA_CLARO = "F2F2F2"


def _cabecalho(ws: Worksheet, linha: int, colunas: list[str]) -> int:
    fonte = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor=_AZUL)
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    for c, titulo in enumerate(colunas, start=1):
        cel = ws.cell(row=linha, column=c, value=titulo)
        cel.font = fonte
        cel.fill = fundo
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = borda
    return linha + 1


def _titulo(ws: Worksheet, linha: int, texto: str, ncols: int) -> int:
    cel = ws.cell(row=linha, column=1, value=texto)
    cel.font = Font(bold=True, size=13, color=_AZUL)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=ncols)
    return linha + 2


def _moeda(ws: Worksheet, linha: int, col: int, valor: Optional[Decimal], negrito: bool = False):
    """Escreve um valor monetário; ``None`` (não informado) vira '-' textual."""
    if valor is None:
        cel = ws.cell(row=linha, column=col, value=NAO_INFORMADO)
        cel.alignment = Alignment(horizontal="right")
    else:
        cel = ws.cell(row=linha, column=col, value=float(valor))
        cel.number_format = _FMT_MOEDA
    if negrito:
        cel.font = Font(bold=True)
    return cel


def _autoajustar(ws: Worksheet, larguras: list[int]) -> None:
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura


def _escrever_cabecalho(ws: Worksheet, cab: Cabecalho | None, ncols: int) -> int:
    """Escreve o cabeçalho de identificação no topo da aba. Retorna a próxima linha."""
    linha = 1
    titulo = ws.cell(row=linha, column=1, value="Retenções Federais sobre NFSe")
    titulo.font = Font(bold=True, size=14, color=_AZUL)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=ncols)
    linha += 1
    for rotulo, valor in linhas_cabecalho(cab):
        cel = ws.cell(row=linha, column=1, value=f"{rotulo}: {valor}")
        cel.font = Font(size=10, color="404040")
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=ncols)
        linha += 1
    return linha + 1  # deixa uma linha em branco antes do relatório


def _aba_sintetico(ws: Worksheet, notas: list[NotaCalculada], cab: Cabecalho | None) -> None:
    colunas = ["Categoria", "Qtd. Notas", "Valor Bruto Total",
               "Total IRRF", "Total CRF", "Total INSS"]
    linha = _escrever_cabecalho(ws, cab, len(colunas))
    linha = _titulo(ws, linha, "Relatório 1 — Totalizador por Tipo de Documento", len(colunas))
    linha = _cabecalho(ws, linha, colunas)

    linhas = agregar_sintetico(notas)
    tot = {"qtd": 0, "bruto": Decimal("0"), "irrf": Decimal("0"),
           "crf": Decimal("0"), "inss": Decimal("0")}
    for ls in linhas:
        ws.cell(row=linha, column=1, value=ls.categoria.rotulo)
        ws.cell(row=linha, column=2, value=ls.qtd_notas)
        _moeda(ws, linha, 3, ls.valor_bruto)
        _moeda(ws, linha, 4, ls.total_irrf)
        _moeda(ws, linha, 5, ls.total_crf)
        _moeda(ws, linha, 6, ls.total_inss)
        tot["qtd"] += ls.qtd_notas
        tot["bruto"] += ls.valor_bruto
        tot["irrf"] += ls.total_irrf
        tot["crf"] += ls.total_crf
        tot["inss"] += ls.total_inss
        linha += 1

    # Linha de total geral.
    fundo = PatternFill("solid", fgColor=_CINZA)
    ws.cell(row=linha, column=1, value="TOTAL GERAL").font = Font(bold=True)
    ws.cell(row=linha, column=2, value=tot["qtd"]).font = Font(bold=True)
    _moeda(ws, linha, 3, tot["bruto"], negrito=True)
    _moeda(ws, linha, 4, tot["irrf"], negrito=True)
    _moeda(ws, linha, 5, tot["crf"], negrito=True)
    _moeda(ws, linha, 6, tot["inss"], negrito=True)
    for c in range(1, len(colunas) + 1):
        ws.cell(row=linha, column=c).fill = fundo
    _autoajustar(ws, [22, 12, 20, 16, 16, 16])


_COLUNAS_ANALITICO = ["Documento", "Número NF", "Data", "Nome / Razão Social", "Qtd. Notas",
                      "Valor Bruto", "IRRF Retido", "CRF Retido", "INSS Retido"]


def _bloco_analitico(ws: Worksheet, linha: int, titulo: str, linhas_dados) -> int:
    linha = _titulo(ws, linha, titulo, len(_COLUNAS_ANALITICO))
    linha = _cabecalho(ws, linha, _COLUNAS_ANALITICO)
    if not linhas_dados:
        ws.cell(row=linha, column=1, value="(sem registros)")
        return linha + 2

    fundo_sub = PatternFill("solid", fgColor=_CINZA_CLARO)
    for la in linhas_dados:
        doc_fmt = mascarar_documento(la.documento, la.tipo)
        # Uma linha por nota (evita listas gigantes numa única célula).
        for item in la.itens_ordenados:
            ws.cell(row=linha, column=1, value=doc_fmt)
            ws.cell(row=linha, column=2, value=item.numero)
            ws.cell(row=linha, column=3, value=item.data_fmt)
            ws.cell(row=linha, column=4, value=la.nome or "-")
            ws.cell(row=linha, column=5, value=1)
            _moeda(ws, linha, 6, item.valor_bruto)
            _moeda(ws, linha, 7, item.irrf)
            _moeda(ws, linha, 8, item.crf)
            _moeda(ws, linha, 9, item.inss)
            linha += 1
        # Subtotal do tomador (só faz sentido destacar quando há mais de 1 nota).
        if la.qtd_notas > 1:
            ws.cell(row=linha, column=4, value=f"Subtotal — {la.nome or doc_fmt}").font = Font(italic=True)
            ws.cell(row=linha, column=5, value=la.qtd_notas).font = Font(bold=True)
            _moeda(ws, linha, 6, la.valor_bruto, negrito=True)
            _moeda(ws, linha, 7, la.total_irrf, negrito=True)
            _moeda(ws, linha, 8, la.total_crf, negrito=True)
            _moeda(ws, linha, 9, la.total_inss, negrito=True)
            for c in range(1, len(_COLUNAS_ANALITICO) + 1):
                ws.cell(row=linha, column=c).fill = fundo_sub
            linha += 1
    return linha + 2


def _aba_analitico(ws: Worksheet, notas: list[NotaCalculada], cab: Cabecalho | None) -> None:
    rel = agregar_analitico(notas)
    linha = _escrever_cabecalho(ws, cab, len(_COLUNAS_ANALITICO))
    linha = _bloco_analitico(ws, linha, "Bloco 1 — Resumo por CNPJ", rel.por_cnpj)
    linha = _bloco_analitico(ws, linha, "Bloco 2 — Resumo por CPF", rel.por_cpf)
    linha = _bloco_analitico(ws, linha, "Bloco 3 — Sem CPF ou CNPJ", rel.sem_documento)

    # Linha de TOTAL GERAL (soma dos três blocos).
    total = rel.total_geral()
    fundo = PatternFill("solid", fgColor=_CINZA)
    ws.cell(row=linha, column=1, value="TOTAL GERAL").font = Font(bold=True)
    ws.cell(row=linha, column=5, value=total.qtd_notas).font = Font(bold=True)
    _moeda(ws, linha, 6, total.valor_bruto, negrito=True)
    _moeda(ws, linha, 7, total.total_irrf, negrito=True)
    _moeda(ws, linha, 8, total.total_crf, negrito=True)
    _moeda(ws, linha, 9, total.total_inss, negrito=True)
    for c in range(1, 10):
        ws.cell(row=linha, column=c).fill = fundo

    _autoajustar(ws, [22, 12, 12, 30, 10, 15, 14, 14, 14])


def _aba_trimestral(ws: Worksheet, notas: list[NotaCalculada], cab: Cabecalho | None) -> None:
    colunas = ["Trimestre", "Qtd. Notas", "Valor Bruto Total",
               "Total IRRF", "Total CRF", "Total INSS"]
    linha = _escrever_cabecalho(ws, cab, len(colunas))
    linha = _titulo(ws, linha, "Relatório 3 — Totalizador Trimestral", len(colunas))
    linha = _cabecalho(ws, linha, colunas)

    linhas = agregar_trimestral(notas)
    tot = {"qtd": 0, "bruto": Decimal("0"), "irrf": Decimal("0"),
           "crf": Decimal("0"), "inss": Decimal("0")}
    if not linhas:
        ws.cell(row=linha, column=1, value="(sem registros)")
        linha += 1
    for lt in linhas:
        ws.cell(row=linha, column=1, value=lt.rotulo)
        ws.cell(row=linha, column=2, value=lt.qtd_notas)
        _moeda(ws, linha, 3, lt.valor_bruto)
        _moeda(ws, linha, 4, lt.total_irrf)
        _moeda(ws, linha, 5, lt.total_crf)
        _moeda(ws, linha, 6, lt.total_inss)
        tot["qtd"] += lt.qtd_notas
        tot["bruto"] += lt.valor_bruto
        tot["irrf"] += lt.total_irrf
        tot["crf"] += lt.total_crf
        tot["inss"] += lt.total_inss
        linha += 1

    fundo = PatternFill("solid", fgColor=_CINZA)
    ws.cell(row=linha, column=1, value="TOTAL GERAL").font = Font(bold=True)
    ws.cell(row=linha, column=2, value=tot["qtd"]).font = Font(bold=True)
    _moeda(ws, linha, 3, tot["bruto"], negrito=True)
    _moeda(ws, linha, 4, tot["irrf"], negrito=True)
    _moeda(ws, linha, 5, tot["crf"], negrito=True)
    _moeda(ws, linha, 6, tot["inss"], negrito=True)
    for c in range(1, len(colunas) + 1):
        ws.cell(row=linha, column=c).fill = fundo
    _autoajustar(ws, [26, 12, 20, 16, 16, 16])


def exportar_excel(
    notas: list[NotaCalculada], caminho: str | Path, cabecalho: Cabecalho | None = None
) -> Path:
    """Gera o arquivo Excel com as três abas de relatório."""
    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sintético (Tipo)"
    _aba_sintetico(ws1, notas, cabecalho)

    ws2 = wb.create_sheet("Analítico (Tomador)")
    _aba_analitico(ws2, notas, cabecalho)

    ws3 = wb.create_sheet("Trimestral")
    _aba_trimestral(ws3, notas, cabecalho)

    wb.save(caminho)
    return caminho
