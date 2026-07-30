from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from retencoes.pipeline import processar

CSV = """numero;documento;nome;valor
1001;11.222.333/0001-81;Empresa Alpha Ltda;10000,00
1002;11.222.333/0001-81;Empresa Alpha Ltda;5000,00
1003;111.444.777-35;Joao Silva;3000,00
1004;;Consumidor Final;800,00
"""


def _csv(tmp_path: Path) -> Path:
    entrada = tmp_path / "notas.csv"
    entrada.write_text(CSV, encoding="utf-8")
    return entrada


def test_fluxo_excel(tmp_path: Path):
    saida = tmp_path / "out" / "retencoes.xlsx"
    gerados, qtd = processar(_csv(tmp_path), saida)

    assert qtd == 4
    assert len(gerados) == 1
    assert gerados[0].exists() and gerados[0].suffix == ".xlsx"

    wb = load_workbook(gerados[0])
    assert "Sintético (Tipo)" in wb.sheetnames
    assert "Analítico (Tomador)" in wb.sheetnames
    assert "Trimestral" in wb.sheetnames


def test_fluxo_pdf(tmp_path: Path):
    saida = tmp_path / "retencoes.pdf"
    gerados, qtd = processar(_csv(tmp_path), saida)

    assert qtd == 4
    assert len(gerados) == 1
    pdf = gerados[0]
    assert pdf.exists() and pdf.suffix == ".pdf"
    # Assinatura de um arquivo PDF válido.
    assert pdf.read_bytes()[:5] == b"%PDF-"


def test_fluxo_ambos(tmp_path: Path):
    saida = tmp_path / "retencoes.xlsx"
    gerados, _ = processar(_csv(tmp_path), saida, formato="ambos")

    sufixos = sorted(g.suffix for g in gerados)
    assert sufixos == [".pdf", ".xlsx"]
    assert all(g.exists() for g in gerados)


def test_diretorio_com_multiplos_arquivos(tmp_path: Path):
    (tmp_path / "a.csv").write_text(CSV, encoding="utf-8")
    (tmp_path / "b.csv").write_text(CSV, encoding="utf-8")
    saida = tmp_path / "retencoes.xlsx"

    _, qtd = processar(tmp_path, saida)
    assert qtd == 8


CSV_SEM_IMPOSTO_INFORMADO = """numero;documento;nome;valor
1001;11.222.333/0001-81;Empresa Alpha Ltda;10000,00
"""

CSV_COM_IMPOSTO_INFORMADO = """numero;documento;nome;valor;irrf;crf;inss
1001;11.222.333/0001-81;Empresa Alpha Ltda;10000,00;150,00;465,00;1100,00
"""


def test_planilha_sem_colunas_de_imposto_nao_calcula_nada(tmp_path: Path):
    entrada = tmp_path / "notas.csv"
    entrada.write_text(CSV_SEM_IMPOSTO_INFORMADO, encoding="utf-8")
    saida = tmp_path / "retencoes.xlsx"

    gerados, qtd = processar(entrada, saida)
    assert qtd == 1
    assert gerados[0].exists()  # não deve lançar erro ao formatar valores None


def test_planilha_com_colunas_de_imposto_usa_valor_informado(tmp_path: Path):
    entrada = tmp_path / "notas.csv"
    entrada.write_text(CSV_COM_IMPOSTO_INFORMADO, encoding="utf-8")
    saida = tmp_path / "retencoes.xlsx"

    from retencoes.ingestao import ler_entrada
    from retencoes.pipeline import calcular_notas

    notas = ler_entrada(entrada)
    calculadas = calcular_notas(notas)
    assert calculadas[0].retencoes.irrf == Decimal("150.00")
    assert calculadas[0].retencoes.crf == Decimal("465.00")
    assert calculadas[0].retencoes.inss == Decimal("1100.00")
